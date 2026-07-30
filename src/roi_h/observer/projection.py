"""Read-only projections over ROI-H's ActiveGraph event stores."""

from __future__ import annotations

import contextlib
import csv
import json
import mimetypes
import re
from datetime import date, datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

from roi_h.observer.activegraph_adapter import ActiveGraphProjectionAdapter

if TYPE_CHECKING:
    from collections.abc import Iterator

_VALID_PROJECT = re.compile(r"^[a-z][a-z0-9_-]{0,63}$")
_VALID_RUN_ID = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,127}$")
_TERMINAL_PHASES = frozenset({"done", "failed", "skipped"})
_TEXT_EXTENSIONS = frozenset(
    {".csv", ".json", ".jsonl", ".log", ".md", ".txt", ".xml", ".yaml", ".yml"}
)
_IMAGE_EXTENSIONS = frozenset({".gif", ".jpeg", ".jpg", ".png", ".webp"})
_MAX_RUNS_PER_DATABASE = 250
_MAX_ARTIFACTS_PER_RUN = 250
_TABLE_PREVIEW_ROWS = 50
_TABLE_PREVIEW_COLUMNS = 20
_TEXT_PREVIEW_BYTES = 131_072
_CURRENT_LAYOUT_VERSION = 4


class ObserverLookupError(LookupError):
    """Raised when an observer resource does not exist."""


def catalog(home: Path) -> dict[str, Any]:
    """Return the projects and environments that have durable databases."""
    home = home.resolve()
    home_config = _read_json(home / "config.json")
    active_project = str(home_config.get("active_project") or home_config.get("project") or "")
    projects: list[dict[str, Any]] = []
    projects_root = home / "projects"
    if not projects_root.is_dir():
        return {"home": str(home), "active_project": active_project, "projects": []}

    for project_root in sorted(path for path in projects_root.iterdir() if path.is_dir()):
        if not _VALID_PROJECT.fullmatch(project_root.name):
            continue
        manifest = project_root / "project.json"
        config = _read_json(manifest if manifest.is_file() else project_root / "config.json")
        environments = []
        for env in ("dev", "prod"):
            current = project_root / "environments" / env / "store" / "activegraph.sqlite"
            legacy = project_root / env / "rpa.sqlite"
            if current.is_file() or legacy.is_file():
                environments.append(env)
        if not environments:
            continue
        projects.append(
            {
                "name": project_root.name,
                "display_name": str(config.get("display_name") or project_root.name),
                "environments": environments,
                "active": project_root.name == active_project,
            }
        )
    return {
        "home": str(home),
        "active_project": active_project,
        "projects": projects,
    }


def list_runs(
    home: Path,
    *,
    project: str | None = None,
    env: str | None = None,
) -> list[dict[str, Any]]:
    """List compact run cards across selected project environments."""
    selected: list[dict[str, Any]] = []
    for workspace in _iter_workspaces(home, project=project, env=env):
        database = workspace["database"]
        adapter = ActiveGraphProjectionAdapter(database)
        for row in adapter.list_run_headers(limit=_MAX_RUNS_PER_DATABASE):
            projection = adapter.project_run(str(row["run_id"]))
            selected.append(
                _run_card(
                    projection,
                    project=str(workspace["project"]),
                    project_display_name=str(workspace["display_name"]),
                    env=str(workspace["env"]),
                    fallback_created_at=str(row["created_at"] or ""),
                    fallback_goal=str(row["goal"] or row["label"] or row["run_id"]),
                )
            )
    return sorted(selected, key=lambda item: str(item["updated_at"]), reverse=True)


def get_run(home: Path, *, project: str, env: str, run_id: str) -> dict[str, Any]:
    """Return the human-facing story for one run."""
    workspace = _workspace(home, project, env)
    adapter = ActiveGraphProjectionAdapter(workspace["database"])
    exists = adapter.run_header(run_id)
    if exists is None:
        msg = f"run not found: {project}/{env}/{run_id}"
        raise ObserverLookupError(msg)
    projection = adapter.project_run(run_id)

    card = _run_card(
        projection,
        project=project,
        project_display_name=str(workspace["display_name"]),
        env=env,
        fallback_created_at=str(exists["created_at"] or ""),
        fallback_goal=str(exists["goal"] or exists["label"] or run_id),
    )
    artifact_root = _artifact_root(workspace, run_id)
    registered_artifacts = _registered_artifacts(projection, artifact_root=artifact_root)
    discovered_artifacts = _discover_artifacts(
        artifact_root,
        registered_artifacts,
    )
    phases = _phase_story(projection, registered_artifacts)
    if discovered_artifacts:
        phases.append(
            {
                "id": "run-artifacts",
                "name": "run_artifacts",
                "title": "Run artifacts",
                "status": "completed",
                "time": _latest_artifact_time(discovered_artifacts),
                "description": "Files stored with this run.",
                "error": None,
                "steps": [],
                "artifacts": discovered_artifacts,
                "synthetic": True,
            }
        )
    return {
        **card,
        "story": phases,
        "technical": _technical_details(projection),
        "artifact_count": len(registered_artifacts) + len(discovered_artifacts),
    }


def preview_artifact(
    home: Path,
    *,
    project: str,
    env: str,
    run_id: str,
    relative_path: str,
) -> dict[str, Any]:
    """Return a bounded browser-friendly preview for an artifact."""
    path, artifact_root = resolve_artifact(
        home,
        project=project,
        env=env,
        run_id=run_id,
        relative_path=relative_path,
    )
    metadata = _artifact_metadata(path, artifact_root)
    extension = path.suffix.lower()
    if extension in {".xlsx", ".xlsm"}:
        return {**metadata, **_preview_workbook(path)}
    if extension == ".csv":
        return {**metadata, **_preview_csv(path)}
    if extension in _TEXT_EXTENSIONS:
        return {**metadata, **_preview_text(path)}
    if extension in _IMAGE_EXTENSIONS or extension == ".pdf":
        return {**metadata, "kind": "native"}
    return {
        **metadata,
        "kind": "unsupported",
        "message": "Preview is not available for this file type.",
    }


def resolve_artifact(
    home: Path,
    *,
    project: str,
    env: str,
    run_id: str,
    relative_path: str,
) -> tuple[Path, Path]:
    """Resolve an artifact and guarantee it remains under the selected run."""
    if not _VALID_RUN_ID.fullmatch(run_id):
        msg = f"invalid run id: {run_id!r}"
        raise ValueError(msg)
    workspace = _workspace(home, project, env)
    artifact_root = _artifact_root(workspace, run_id).resolve()
    path = (artifact_root / relative_path).resolve()
    try:
        path.relative_to(artifact_root)
    except ValueError as exc:
        msg = "artifact path escapes the selected run"
        raise ValueError(msg) from exc
    if not path.is_file():
        msg = f"artifact not found: {relative_path}"
        raise ObserverLookupError(msg)
    return path, artifact_root


def _iter_workspaces(
    home: Path,
    *,
    project: str | None,
    env: str | None,
) -> Iterator[dict[str, Any]]:
    current_catalog = catalog(home)
    for item in current_catalog["projects"]:
        project_name = str(item["name"])
        if project is not None and project_name != project:
            continue
        for environment in item["environments"]:
            if env is not None and environment != env:
                continue
            yield _workspace(home, project_name, str(environment))


def _workspace(home: Path, project: str, env: str) -> dict[str, Any]:
    if not _VALID_PROJECT.fullmatch(project):
        msg = f"invalid project: {project!r}"
        raise ValueError(msg)
    if env not in {"dev", "prod"}:
        msg = f"invalid environment: {env!r}"
        raise ValueError(msg)
    project_root = home.resolve() / "projects" / project
    manifest = project_root / "project.json"
    current = manifest.is_file()
    config = _read_json(manifest if current else project_root / "config.json")
    database = (
        project_root / "environments" / env / "store" / "activegraph.sqlite"
        if current
        else project_root / env / "rpa.sqlite"
    )
    if not database.is_file():
        msg = f"database not found: {project}/{env}"
        raise ObserverLookupError(msg)
    return {
        "project": project,
        "display_name": str(config.get("display_name") or project),
        "env": env,
        "database": database.resolve(),
        "runs": (
            (project_root / "environments" / env / "runs").resolve()
            if current
            else (project_root / env / "artifacts").resolve()
        ),
        "layout_version": 4 if current else 3,
    }


def _artifact_root(workspace: dict[str, Any], run_id: str) -> Path:
    root = Path(workspace["runs"]) / run_id
    if workspace.get("layout_version") == _CURRENT_LAYOUT_VERSION:
        root /= "artifacts"
    return root


def _run_card(  # noqa: PLR0913
    projection: dict[str, Any],
    *,
    project: str,
    project_display_name: str,
    env: str,
    fallback_created_at: str,
    fallback_goal: str,
) -> dict[str, Any]:
    objects = projection["objects"]
    run_object = next((item for item in objects if item["type"] == "rpa.run"), None)
    run_data = dict(run_object["data"]) if run_object else {}
    phases = [item for item in objects if item["type"] == "rpa.phase"]
    steps = [item for item in objects if item["type"] == "rpa.step"]
    approvals = [item for item in objects if item["type"] == "rpa.approval"]
    invocations = [item for item in objects if item["type"] == "rpa.invocation"]
    failed_step = next(
        (item for item in reversed(steps) if str(item["data"].get("status") or "") == "error"),
        None,
    )
    failed_phase = next(
        (item for item in reversed(phases) if str(item["data"].get("status") or "") == "failed"),
        None,
    )
    pending_approval = next(
        (item for item in approvals if str(item["data"].get("status") or "") == "pending"),
        None,
    )
    unknown_invocation = next(
        (
            item
            for item in invocations
            if str(item["data"].get("status") or "") == "outcome_unknown"
        ),
        None,
    )
    status = _derive_run_status(
        run_data,
        phases=phases,
        failed_step=failed_step,
        failed_phase=failed_phase,
        pending_approval=pending_approval,
        unknown_invocation=unknown_invocation,
    )
    if status == "idle" and steps:
        status = "completed"
    title = str(
        run_data.get("automation_name")
        or run_data.get("goal")
        or fallback_goal
        or projection["run_id"]
    )
    created_at = str(projection["first_timestamp"] or fallback_created_at)
    updated_at = str(projection["last_timestamp"] or created_at)
    duration_seconds = _duration_seconds(created_at, updated_at)
    return {
        "run_id": projection["run_id"],
        "project": project,
        "project_display_name": project_display_name,
        "env": env,
        "title": title,
        "goal": str(run_data.get("goal") or title),
        "status": status,
        "attention": status in {"failed", "unknown", "approval"},
        "summary": _run_summary(
            status,
            run_data=run_data,
            phases=phases,
            steps=steps,
            failed_step=failed_step,
            failed_phase=failed_phase,
        ),
        "created_at": created_at,
        "updated_at": updated_at,
        "duration_seconds": duration_seconds,
        "phase_count": len(phases),
        "step_count": len(steps),
        "current_phase": run_data.get("current_phase"),
    }


def _derive_run_status(  # noqa: PLR0911, PLR0913
    run_data: dict[str, Any],
    *,
    phases: list[dict[str, Any]],
    failed_step: dict[str, Any] | None,
    failed_phase: dict[str, Any] | None,
    pending_approval: dict[str, Any] | None,
    unknown_invocation: dict[str, Any] | None,
) -> str:
    if unknown_invocation is not None:
        return "unknown"
    if pending_approval is not None:
        return "approval"
    if failed_step is not None or failed_phase is not None:
        return "failed"
    recorded = str(run_data.get("status") or "").lower()
    if recorded in {"failed", "cancelled", "completed"}:
        return recorded
    if phases and all(str(item["data"].get("status") or "") in _TERMINAL_PHASES for item in phases):
        return "completed"
    if run_data.get("current_phase") or any(
        str(item["data"].get("status") or "") == "open" for item in phases
    ):
        return "running"
    return "idle"


def _run_summary(  # noqa: PLR0911, PLR0913
    status: str,
    *,
    run_data: dict[str, Any],
    phases: list[dict[str, Any]],
    steps: list[dict[str, Any]],
    failed_step: dict[str, Any] | None,
    failed_phase: dict[str, Any] | None,
) -> str:
    if status == "failed":
        failed = failed_phase or failed_step
        if failed is None:
            return "A recorded step failed"
        phase_name = str(failed["data"].get("phase") or failed["data"].get("name") or "a step")
        return f"Failed during {_display_name(phase_name)}"
    if status == "unknown":
        return "The outcome of a write could not be confirmed"
    if status == "approval":
        return "Waiting for operator approval"
    if status == "completed":
        if phases:
            return f"{len(phases)} phases completed"
        return f"{len(steps)} steps completed"
    if status == "cancelled":
        return str(run_data.get("cancel_reason") or "Run cancelled")
    if status == "running":
        current = str(run_data.get("current_phase") or "Run")
        return f"{_display_name(current)} is in progress"
    return "Run recorded"


def _phase_story(
    projection: dict[str, Any],
    artifacts: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    objects = projection["objects"]
    run_object = next((item for item in objects if item["type"] == "rpa.run"), None)
    run_data = dict(run_object["data"]) if run_object else {}
    phases = sorted(
        (item for item in objects if item["type"] == "rpa.phase"),
        key=lambda item: (int(item["data"].get("index") or 0), str(item["created_at"])),
    )
    steps = [item for item in objects if item["type"] == "rpa.step"]
    story: list[dict[str, Any]] = []
    materialized_names: set[str] = set()
    for phase in phases:
        phase_data = phase["data"]
        phase_id = str(phase["id"])
        name = str(phase_data.get("name") or "phase")
        materialized_names.add(name)
        phase_steps = [
            _step_view(item)
            for item in steps
            if item["data"].get("phase_id") == phase_id
            or (not item["data"].get("phase_id") and str(item["data"].get("phase") or "") == name)
        ]
        phase_artifacts = [
            item
            for item in artifacts
            if item.get("phase_id") == phase_id or item.get("phase") == name
        ]
        status = _derive_phase_status(phase_data, phase_steps)
        story.append(
            {
                "id": phase_id,
                "name": name,
                "title": _display_name(name),
                "status": status,
                "time": _time_label(str(phase["created_at"])),
                "description": _phase_description(phase_data, phase_steps, status),
                "error": phase_data.get("error") or _first_step_error(phase_steps),
                "steps": phase_steps,
                "artifacts": phase_artifacts,
                "synthetic": False,
            }
        )

    for raw in run_data.get("phase_plan") or []:
        if not isinstance(raw, dict):
            continue
        name = str(raw.get("name") or "")
        if not name or name in materialized_names:
            continue
        story.append(
            {
                "id": f"planned-{name}",
                "name": name,
                "title": _display_name(name),
                "status": "not_started",
                "time": "",
                "description": str(raw.get("description") or "This phase has not started."),
                "error": None,
                "steps": [],
                "artifacts": [],
                "synthetic": True,
            }
        )

    unphased_steps = [
        _step_view(item)
        for item in steps
        if not item["data"].get("phase_id") and not item["data"].get("phase")
    ]
    if unphased_steps:
        story.insert(
            0,
            {
                "id": "unphased-activity",
                "name": "activity",
                "title": "Activity",
                "status": _derive_phase_status({}, unphased_steps),
                "time": unphased_steps[0]["time"],
                "description": f"{len(unphased_steps)} recorded steps.",
                "error": _first_step_error(unphased_steps),
                "steps": unphased_steps,
                "artifacts": [
                    item for item in artifacts if not item.get("phase_id") and not item.get("phase")
                ],
                "synthetic": True,
            },
        )
    return story


def _derive_phase_status(phase_data: dict[str, Any], steps: list[dict[str, Any]]) -> str:
    if any(item["status"] == "outcome_unknown" for item in steps):
        return "unknown"
    if any(item["status"] == "error" for item in steps):
        return "failed"
    status = str(phase_data.get("status") or "")
    return {
        "done": "completed",
        "failed": "failed",
        "skipped": "skipped",
        "open": "running",
    }.get(status, "completed" if steps else "idle")


def _phase_description(  # noqa: PLR0911
    phase_data: dict[str, Any],
    steps: list[dict[str, Any]],
    status: str,
) -> str:
    description = str(phase_data.get("description") or "").strip()
    if description:
        return description
    summary = phase_data.get("summary")
    if isinstance(summary, dict):
        for value in summary.values():
            if isinstance(value, str) and value.strip():
                return value.strip()
    if status == "failed":
        return "This phase stopped before it could complete."
    if status == "running":
        return "Work is currently in progress."
    if status == "skipped":
        return "This phase was skipped."
    if steps:
        suffix = "" if len(steps) == 1 else "s"
        return f"{len(steps)} step{suffix} completed."
    return "Phase completed."


def _step_view(item: dict[str, Any]) -> dict[str, Any]:
    data = item["data"]
    return {
        "id": item["id"],
        "name": str(data.get("name") or f"{data.get('skill')}.{data.get('tool')}"),
        "skill": str(data.get("skill") or ""),
        "tool": str(data.get("tool") or ""),
        "status": str(data.get("status") or "unknown"),
        "error": data.get("error"),
        "duration_seconds": data.get("duration_seconds"),
        "time": _time_label(str(data.get("started_at") or item["created_at"])),
        "invocation_id": data.get("invocation_id"),
        "attempt": data.get("attempt"),
        "args": data.get("args") if isinstance(data.get("args"), dict) else {},
        "output": data.get("output") if isinstance(data.get("output"), dict) else {},
    }


def _registered_artifacts(
    projection: dict[str, Any],
    *,
    artifact_root: Path,
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    root = artifact_root.resolve()
    for item in projection["objects"]:
        if item["type"] != "rpa.artifact":
            continue
        data = item["data"]
        raw_path = Path(str(data.get("path") or ""))
        artifact_id = str(data.get("artifact_id") or "")
        identity_files = (
            [
                path
                for path in root.glob(f"{artifact_id}--*")
                if path.is_file() and not path.is_symlink()
            ]
            if artifact_id
            else []
        )
        candidates = [
            *identity_files,
            raw_path,
            root / str(data.get("name") or ""),
        ]
        resolved = next(
            (
                candidate.resolve()
                for candidate in candidates
                if candidate.is_file() and candidate.resolve().is_relative_to(root)
            ),
            None,
        )
        if resolved is None:
            continue
        metadata = _artifact_metadata(resolved, root)
        metadata.update(
            {
                "name": str(data.get("name") or metadata["name"]),
                "phase": data.get("phase"),
                "phase_id": data.get("phase_id"),
                "artifact_id": data.get("artifact_id"),
                "uri": data.get("uri"),
                "sha256": data.get("sha256"),
                "registered": True,
            }
        )
        items.append(metadata)
    return items


def _discover_artifacts(
    artifact_root: Path,
    registered: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if not artifact_root.is_dir():
        return []
    known = {str(item["relative_path"]) for item in registered}
    items: list[dict[str, Any]] = []
    root = artifact_root.resolve()
    for path in sorted(root.rglob("*")):
        if len(items) >= _MAX_ARTIFACTS_PER_RUN:
            break
        if not path.is_file() or path.is_symlink():
            continue
        relative = path.relative_to(root)
        if "phases" in relative.parts or relative.as_posix() in known:
            continue
        items.append(
            {
                **_artifact_metadata(path, root),
                "phase": None,
                "phase_id": None,
                "sha256": None,
                "registered": False,
            }
        )
    return items


def _artifact_metadata(path: Path, artifact_root: Path) -> dict[str, Any]:
    stat = path.stat()
    extension = path.suffix.lower()
    if extension in {".xlsx", ".xlsm", ".csv"}:
        preview_kind = "table"
    elif extension in _TEXT_EXTENSIONS:
        preview_kind = "text"
    elif extension in _IMAGE_EXTENSIONS:
        preview_kind = "image"
    elif extension == ".pdf":
        preview_kind = "pdf"
    else:
        preview_kind = "unsupported"
    return {
        "name": path.name,
        "relative_path": path.relative_to(artifact_root).as_posix(),
        "bytes": stat.st_size,
        "modified_at": datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(),
        "mime_type": mimetypes.guess_type(path.name)[0] or "application/octet-stream",
        "preview_kind": preview_kind,
    }


def _preview_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = [
            [_json_value(value) for value in row]
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=min(worksheet.max_row, _TABLE_PREVIEW_ROWS),
                max_col=min(worksheet.max_column, _TABLE_PREVIEW_COLUMNS),
                values_only=True,
            )
        ]
        rows = _trim_table(rows)
        return {
            "kind": "table",
            "sheet": worksheet.title,
            "sheets": list(workbook.sheetnames),
            "rows": rows,
            "total_rows": worksheet.max_row,
            "total_columns": worksheet.max_column,
            "truncated": (
                worksheet.max_row > _TABLE_PREVIEW_ROWS
                or worksheet.max_column > _TABLE_PREVIEW_COLUMNS
            ),
        }
    finally:
        workbook.close()


def _preview_csv(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        rows = [
            row[:_TABLE_PREVIEW_COLUMNS]
            for _, row in zip(range(_TABLE_PREVIEW_ROWS), csv.reader(handle), strict=False)
        ]
    return {
        "kind": "table",
        "sheet": None,
        "sheets": [],
        "rows": _trim_table(rows),
        "total_rows": None,
        "total_columns": max((len(row) for row in rows), default=0),
        "truncated": len(rows) == _TABLE_PREVIEW_ROWS,
    }


def _preview_text(path: Path) -> dict[str, Any]:
    content = path.read_bytes()[:_TEXT_PREVIEW_BYTES].decode("utf-8", errors="replace")
    if path.suffix.lower() == ".json":
        with contextlib.suppress(json.JSONDecodeError):
            content = json.dumps(json.loads(content), indent=2, ensure_ascii=False)
    return {
        "kind": "text",
        "content": content,
        "truncated": path.stat().st_size > _TEXT_PREVIEW_BYTES,
    }


def _technical_details(projection: dict[str, Any]) -> dict[str, Any]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in projection["objects"]:
        grouped.setdefault(str(item["type"]), []).append(
            {
                "id": item["id"],
                "data": item["data"],
                "created_at": item["created_at"],
                "updated_at": item["updated_at"],
            }
        )
    return {
        "run_id": projection["run_id"],
        "objects": grouped,
        "first_event_at": projection["first_timestamp"],
        "last_event_at": projection["last_timestamp"],
    }


def _latest_artifact_time(artifacts: list[dict[str, Any]]) -> str:
    if not artifacts:
        return ""
    latest = max(str(item["modified_at"]) for item in artifacts)
    return _time_label(latest)


def _first_step_error(steps: list[dict[str, Any]]) -> str | None:
    return next((str(item["error"]) for item in steps if item.get("error")), None)


def _duration_seconds(start: str, end: str) -> float | None:
    start_dt = _parse_datetime(start)
    end_dt = _parse_datetime(end)
    if start_dt is None or end_dt is None:
        return None
    return max(0.0, (end_dt - start_dt).total_seconds())


def _parse_datetime(value: str) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def _time_label(value: str) -> str:
    parsed = _parse_datetime(value)
    if parsed is None:
        return ""
    return parsed.astimezone().strftime("%H:%M")


def _display_name(value: str) -> str:
    return value.replace("-", " ").replace("_", " ").strip().capitalize()


def _read_json(path: Path) -> dict[str, Any]:
    if not path.is_file():
        return {}
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return value if isinstance(value, dict) else {}


def _json_value(value: object) -> object:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _trim_table(rows: list[list[Any]]) -> list[list[Any]]:
    while rows and all(value in {None, ""} for value in rows[-1]):
        rows.pop()
    max_index = 0
    for row in rows:
        for index, value in enumerate(row, start=1):
            if value not in {None, ""}:
                max_index = max(max_index, index)
    return [row[:max_index] for row in rows]


__all__ = [
    "ObserverLookupError",
    "catalog",
    "get_run",
    "list_runs",
    "preview_artifact",
    "resolve_artifact",
]
