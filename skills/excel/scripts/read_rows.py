from __future__ import annotations
import csv
from pathlib import Path
from pydantic import BaseModel, Field
TOOL_ID="read_rows"; DESCRIPTION="Read rows from xlsx or csv as list of dicts."; DETERMINISTIC=True; REQUIRES_APPROVAL=False
FILESYSTEM_ROOTS=("project:reference:read","run:input:read","run:work:read-write","run:output:read-write","artifact:read","automation:read")
TOOL_EFFECT = 'read'
IDEMPOTENCY = 'none'
ALLOW_IN_PROD = True
TIMEOUT_SECONDS = 120.0
SECRET_NAMES = ()
NETWORK_HOSTS = ()


class Input(BaseModel):
    path: str
    sheet: str = ""
    max_rows: int = Field(default=0, description="0=all")
class Output(BaseModel):
    ok: bool = True
    path: str
    headers: list[str]
    rows: list[dict]
    count: int
def run(args: Input) -> Output:
    p = Path(args.path).expanduser()
    if p.suffix.lower() in {".csv", ".tsv"}:
        delim = "\t" if p.suffix.lower()==".tsv" else ","
        with p.open(encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f, delimiter=delim)
            headers = list(reader.fieldnames or [])
            rows = [dict(r) for r in reader]
    else:
        from openpyxl import load_workbook
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb[args.sheet] if args.sheet else wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(it)]
        rows = []
        for raw in it:
            if raw is None or all(c is None or str(c).strip()=="" for c in raw):
                continue
            rows.append({headers[i]: ("" if raw[i] is None else str(raw[i]).strip()) for i in range(len(headers)) if headers[i]})
        wb.close()
    if args.max_rows > 0:
        rows = rows[: args.max_rows]
    return Output(ok=True, path=str(p.resolve()), headers=headers, rows=rows, count=len(rows))
